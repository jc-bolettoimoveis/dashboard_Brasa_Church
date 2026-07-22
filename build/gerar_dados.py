#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o docs/dados.json do painel da Brasa Church LENDO A PLANILHA DO GOOGLE.
Roda sozinho no GitHub Actions (semanal + botao manual). Voce so preenche a planilha.

Passos internos:
  1. Baixa a planilha do Google como .xlsx (precisa estar como "qualquer pessoa com o link: leitor").
  2. Le as abas de cada culto (mesma estrutura da planilha atual) e soma os totais por domingo.
  3. Busca o clima hora a hora em Porto Alegre (Open-Meteo, gratis) para cada domingo.
  4. Junta feriados / Copa / local (arquivo build/fatores.json, editavel).
  5. Escreve docs/dados.json — que o index.html ja consome.
"""
import os, re, json, io, datetime, sys
import requests
import openpyxl

# ------- CONFIG -------
SHEET_ID = os.environ.get("SHEET_ID", "1Bkl6jAmyccn70lBGsV-IJh5-qtDtoTA5nOuP4Kbvd_4")
ANO      = int(os.environ.get("ANO", "2026"))
LAT, LON = -30.0331, -51.23
AQUI     = os.path.dirname(os.path.abspath(__file__))
RAIZ     = os.path.dirname(AQUI)
# Onde gravar o dados.json: na mesma pasta do index.html.
# No repositorio da Brasa o index.html esta na raiz, entao gravamos na raiz.
SAIDA    = os.path.join(RAIZ, "dados.json")
HRS      = ['9', '11', '16', '18', '20']
MES_ABREV = ['', 'Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

# ------- 1. baixar a planilha -------
def baixar_xlsx():
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    r = requests.get(url, timeout=90)
    r.raise_for_status()
    if b"<html" in r.content[:200].lower():
        raise SystemExit("A planilha nao esta acessivel publicamente. "
                         "No Google Sheets: Compartilhar > Acesso geral > 'Qualquer pessoa com o link' = Leitor.")
    return io.BytesIO(r.content)

# ------- 2. ler as abas -------
def parse_teatro(ws):
    for r in ws.iter_rows(values_only=True):
        if r and isinstance(r[0], str) and r[0].strip().upper().startswith('TOTAL'):
            for v in reversed(r):
                if isinstance(v, (int, float)): return int(v)
    return None

def parse_centro(ws):
    servicos, cur = {}, None
    for r in ws.iter_rows(values_only=True):
        joined = ' '.join(str(c) for c in r if c)
        m = re.search(r'culto das (\d+)\s*hs', joined.lower())
        if m: cur = int(m.group(1))
        for c in r:
            if isinstance(c, str) and c.strip().lower() == 'total':
                for v in reversed(r):
                    if isinstance(v, (int, float)):
                        if cur is not None: servicos[cur] = int(v)
                        break
    return servicos

def to_date(tok):
    tok = tok.strip()
    if len(tok) == 3:  d, m = int(tok[:2]), int(tok[2:])
    else:              tok = tok.zfill(4); d, m = int(tok[:2]), int(tok[2:])
    return datetime.date(ANO, m, d)

def ler_planilha(xls):
    wb = openpyxl.load_workbook(xls, data_only=True)
    skip = {'MODELO  CENTRO', 'MODELO TEATRO', 'GERAL', 'Página5', 'PÃ¡gina5'}
    teatro, centro = {}, {}
    for name in wb.sheetnames:
        nm = name.strip()
        if nm in skip or nm.lower().startswith('cópia') or nm.lower().startswith('copia'):
            continue
        ws = wb[name]
        if re.search(r'\d{3,4}\s*[-\s].*\d+\s*[hH]', nm):        # aba de horario (Teatro)
            dm = re.match(r'^\s*(\d{3,4})', nm); hm = re.search(r'(\d{1,2})\s*[hH]', nm)
            if dm and hm:
                try: d = to_date(dm.group(1))
                except ValueError: continue
                teatro.setdefault(d, {})[str(int(hm.group(1)))] = parse_teatro(ws)
        elif re.match(r'^\d{3,4}$', nm):                          # aba unica (Centro de Eventos)
            try: d = to_date(nm)
            except ValueError: continue
            centro[d] = {str(k): v for k, v in parse_centro(ws).items()}
    base = {}
    for d, s in teatro.items(): base[d] = {'venue': 'Teatro', 'services': {h: v for h, v in s.items() if v}}
    for d, s in centro.items(): base[d] = {'venue': 'Centro', 'services': {h: v for h, v in s.items() if v}}
    out = []
    for d in sorted(base):
        serv = base[d]['services']
        total = sum(v for v in serv.values() if v)
        if total <= 0: continue
        out.append({'date': d.isoformat(), 'venue': base[d]['venue'], 'services': serv, 'total': total})
    return out

# ------- 3. clima -------
def cond(p, c):
    if p >= 0.4: return 'Chuva'
    if p >= 0.1: return 'Garoa'
    if c < 30:  return 'Sol'
    if c < 70:  return 'Parcial'
    return 'Nublado'

def buscar_clima(datas):
    if not datas: return {}
    mn, mx = datas[0], datas[-1]
    url = ("https://archive-api.open-meteo.com/v1/archive?latitude=%s&longitude=%s"
           "&start_date=%s&end_date=%s"
           "&hourly=temperature_2m,precipitation,cloudcover,apparent_temperature,windspeed_10m"
           "&timezone=America%%2FSao_Paulo" % (LAT, LON, mn, mx))
    try:
        j = requests.get(url, timeout=90).json()
    except Exception as e:
        print("Aviso: clima indisponivel (%s). Segue sem clima." % e); return {}
    t = j['hourly']['time']; idx = {t[i]: i for i in range(len(t))}
    H = j['hourly']
    wx = {}
    for d in datas:
        wx[d] = {}
        for h in HRS:
            key = "%sT%02d:00" % (d, int(h))
            i = idx.get(key)
            if i is None or H['temperature_2m'][i] is None:
                wx[d][h] = {"temp": None, "app": None, "wind": None, "precip": 0, "cloud": 0, "cond": "Nublado"}
            else:
                p = H['precipitation'][i] or 0; c = H['cloudcover'][i] or 0
                wx[d][h] = {"temp": round(H['temperature_2m'][i], 1),
                            "app": round(H['apparent_temperature'][i], 1),
                            "wind": round(H['windspeed_10m'][i], 1),
                            "precip": p, "cloud": round(c), "cond": cond(p, c)}
    return wx

# ------- 4. montar dados.json -------
def montar(base, wx, fatores, kids, funil):
    F = fatores
    dados = []
    for r in base:
        d = r['date']; y, m, dd = d.split('-')
        row = {
            "date": d, "label": "%s/%s" % (dd, m), "mes": MES_ABREV[int(m)],
            "venue": "Centro" if d in F.get("centro", []) else "Teatro",
            "total": r['total'], "services": r['services'],
            "holiday": F.get("feriados", {}).get(d, ""),
            "wc": F.get("copa", {}).get(d, ""),
            "note": F.get("notas", {}).get(d, ""),
            "wx": wx.get(d, {})
        }
        k = kids.get(d)
        if k and k.get('total', 0) > 0:
            row["kids"] = k['total']; row["kidsHour"] = k.get('byhour', {}); row["kidsVis"] = k.get('vis', 0)
            row["familia"] = round(100 * k['total'] / r['total'], 1) if r['total'] else None
        fu = funil.get(d)
        if fu:
            if fu.get('pv') is not None: row["pv"] = fu['pv']   # primeira vez (Boas-vindas)
            if fu.get('aj') is not None: row["aj"] = fu['aj']   # aceitou Jesus
        dados.append(row)
    return dados

def main():
    print("Baixando a planilha do Google...")
    base = ler_planilha(baixar_xlsx())
    print("Domingos encontrados: %d" % len(base))
    if not base:
        raise SystemExit("Nenhum domingo de %d encontrado. Confira a planilha." % ANO)
    datas = [r['date'] for r in base]
    print("Buscando o clima...")
    wx = buscar_clima(datas)
    fatores = {}
    fpath = os.path.join(AQUI, "fatores.json")
    if os.path.exists(fpath):
        fatores = json.load(open(fpath, encoding="utf-8"))
    kids = {}
    kpath = os.path.join(AQUI, "kids.json")   # presenca do Kids (contagens, sem dado pessoal)
    if os.path.exists(kpath):
        kids = json.load(open(kpath, encoding="utf-8"))
    funil = {}
    fupath = os.path.join(AQUI, "funil.json")  # primeira vez / aceitou Jesus (so quantidades)
    if os.path.exists(fupath):
        funil = json.load(open(fupath, encoding="utf-8"))
    dados = montar(base, wx, fatores, kids, funil)
    with open(SAIDA, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=1)
    print("OK: dados.json com %d domingos, total %d pessoas." %
          (len(dados), sum(d['total'] for d in dados)))

if __name__ == "__main__":
    main()
